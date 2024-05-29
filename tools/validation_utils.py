import openpyxl
import os
import torch
import numpy as np

def create_or_modify_excel(file_path, data_path, threshold, model, data):
    """
    Create or modify an Excel file to append detection results.
    
    Args:
        file_path (str): Path to the Excel file.
        data_path (str): Path to the data directory.
        threshold (float): IoU threshold for detection.
        model (str): Model name.
        data (list): List of detection results.
    """
    parsed_data = parse_path_and_data(data_path, threshold, model, data)
    
    if os.path.exists(file_path):
        # If the file already exists, load it and append data
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # Check if the worksheet is empty and add column names if necessary
        if ws.max_row == 0:
            ws.append(['Sensor', 'Condition', 'Detector', 'IoU Threshold', 'Number of cases', 'Successful detections'])
        ws.append(parsed_data)
    else:
        # If the file doesn't exist, create a new one and add data
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Sensor', 'Condition', 'Detector', 'IoU Threshold', 'Number of cases', 'Successful detections'])
        ws.append(parsed_data)
    
    wb.save(file_path)

def create_or_modify_excel_recall(file_path, data_path, threshold, model, data, budget):
    """
    Create or modify an Excel file to append recall results.
    
    Args:
        file_path (str): Path to the Excel file.
        data_path (str): Path to the data directory.
        threshold (float): IoU threshold for detection.
        model (str): Model name.
        data (list): List of detection results.
        budget (int): Budget for the recall calculation.
    """
    parsed_data = parse_path_and_data_recall(data_path, threshold, model, data, budget)
    
    if os.path.exists(file_path):
        # If the file already exists, load it and append data
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # Check if the worksheet is empty and add column names if necessary
        if ws.max_row == 0:
            ws.append(['Sensor', 'Condition', 'Detector', 'IoU Threshold', 'Budget', 'Recall'])
        ws.append(parsed_data)
    else:
        # If the file doesn't exist, create a new one and add data
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Sensor', 'Condition', 'Detector', 'IoU Threshold', 'Budget', 'Recall'])
        ws.append(parsed_data)
    
    wb.save(file_path)

def create_or_modify_excel_recall_distance(file_path, data_path, threshold, model, data, budget, distances):
    """
    Create or modify an Excel file to append recall results.
    
    Args:
        file_path (str): Path to the Excel file.
        data_path (str): Path to the data directory.
        threshold (float): IoU threshold for detection.
        model (str): Model name.
        data (list): List of detection results.
        budget (int): Budget for the recall calculation.
    """
    parsed_data = parse_path_and_data_distance(data_path, threshold, model, data, budget, distances)
    
    if os.path.exists(file_path):
        # If the file already exists, load it and append data
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # Check if the worksheet is empty and add column names if necessary
        if ws.max_row == 0:
            ws.append(['Sensor', 'Condition', 'Detector', 'IoU Threshold', 'Budget', 'Distance', 'Recall'])
        ws.append(parsed_data)
    else:
        # If the file doesn't exist, create a new one and add data
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Sensor', 'Condition', 'Detector', 'IoU Threshold', 'Budget', 'Distance', 'Recall'])
        ws.append(parsed_data)
    
    wb.save(file_path)

def create_or_modify_excel_generic(scores, paths, detector, file_path="genetic_results.xlsx"):
    sensors = ["HDL-64E", "VLP-16"]
    conditions = ["clear", "light", "moderate", "heavy", "extreme"]
    thresholds = [0.0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9]
    results = zip(scores, paths)
    for sensor in sensors:
        sensor_results = [(iou, path) for iou, path in results if sensor in path]
        for condition in conditions:
            condition_results = [(iou, path) for iou, path in sensor_results if condition in path]
            if len(condition_results) > 0:
                for threshold in thresholds:
                    total = len(condition_results)
                    positive = sum(1 for x, _ in condition_results if x > threshold)
                    entry = [sensor, condition, detector, "genetic", threshold, 200, total, positive, positive / total]

                    if os.path.exists(file_path):
                        # If the file already exists, load it and append data
                        wb = openpyxl.load_workbook(file_path)
                        ws = wb.active
                        # Check if the worksheet is empty and add column names if necessary
                        if ws.max_row == 0:
                            ws.append(['Sensor', 'Condition', 'Detector', 'Attack type', 'IoU Threshold', 'Budget', 'Total', 'Positive', 'Recall'])
                        exists = False
                        for row in ws.iter_rows(min_row=2):
                            if row[1] == sensor and row[2] == condition and row[5] == threshold:
                                exists = True
                                row[7] += total
                                row[8] += positive
                                row[9] == row[7] / row[8]
                        if not exists:
                            ws.append(entry)
                    else:
                        # If the file doesn't exist, create a new one and add data
                        wb = openpyxl.Workbook()
                        ws = wb.active
                        ws.append(['Sensor', 'Condition', 'Detector', 'Attack type', 'IoU Threshold', 'Budget', 'Total', 'Positive', 'Recall'])
                        ws.append(entry)
                    
                    wb.save(file_path)




def parse_path_and_data(data_path, threshold, model, data):
    """
    Parse the data path and detection results for appending to Excel.
    
    Args:
        data_path (str): Path to the data directory.
        threshold (float): IoU threshold for detection.
        model (str): Model name.
        data (list): List of detection results.
    
    Returns:
        list: Parsed data.
    """
    folders = data_path.split(os.path.sep)
    output = []
    
    # Determine the sensor type
    if 'HDL-64E' in folders:
        output.append('HDL-64E')
    else:
        output.append('VLP-16')
    
    # Determine the condition type
    if 'clear' in folders:
        output.append('clear')
    elif 'light' in folders:
        output.append('light')
    elif 'moderate' in folders:
        output.append('moderate')
    elif 'heavy' in folders:
        output.append('heavy')
    elif 'extreme' in folders:
        output.append('extreme')
    
    output.append(model)
    output.append(threshold)
    output.append(len(data))
    output.append(sum(1 for x in data if x > threshold))
    
    return output

def parse_path_and_data_recall(data_path, threshold, model, data, budget):
    """
    Parse the data path and recall results for appending to Excel.
    
    Args:
        data_path (str): Path to the data directory.
        threshold (float): IoU threshold for detection.
        model (str): Model name.
        data (list): List of detection results.
        budget (int): Budget for the recall calculation.
    
    Returns:
        list: Parsed data.
    """
    folders = data_path.split(os.path.sep)
    output = []
    
    # Determine the sensor type
    if 'HDL-64E' in folders:
        output.append('HDL-64E')
    else:
        output.append('VLP-16')
    
    # Determine the condition type
    if 'clear' in folders:
        output.append('clear')
    elif 'light' in folders:
        output.append('light')
    elif 'moderate' in folders:
        output.append('moderate')
    elif 'heavy' in folders:
        output.append('heavy')
    elif 'extreme' in folders:
        output.append('extreme')
    
    output.append(model)
    output.append(threshold)
    output.append(budget)
    output.append(sum(1 for x in data if x > threshold) / len(data))
    
    return output

def parse_path_and_data_distance(data_path, threshold, model, data, budget, distance):
    """
    Parse the data path and recall results for appending to Excel.
    
    Args:
        data_path (str): Path to the data directory.
        threshold (float): IoU threshold for detection.
        model (str): Model name.
        data (list): List of detection results.
        budget (int): Budget for the recall calculation.
    
    Returns:
        list: Parsed data.
    """
    folders = data_path.split(os.path.sep)
    output = []
    
    # Determine the sensor type
    if 'HDL-64E' in folders:
        output.append('HDL-64E')
    else:
        output.append('VLP-16')
    
    # Determine the condition type
    if 'clear' in folders:
        output.append('clear')
    elif 'light' in folders:
        output.append('light')
    elif 'moderate' in folders:
        output.append('moderate')
    elif 'heavy' in folders:
        output.append('heavy')
    elif 'extreme' in folders:
        output.append('extreme')
    output.append(model)
    output.append(threshold)
    output.append(budget)
    output.append(distance)
    output.append(sum(1 for x in data if x > threshold) / len(data))
    
    return output

def find_verification_files(condition_path):
    verification_files = []
    for root, dirs, files in os.walk(condition_path):
        for file in files:
            if file.endswith('_verification.npy'):
                verification_files.append(os.path.join(root, file))
    return verification_files

def calculate_minimum_distance_for_points(points):
    if len(points) == 0:
        return 99999
    xyz_coordinates = points[:, :3]
    distances = np.sqrt(np.sum(xyz_coordinates**2, axis=1))
    return np.min(distances)

def get_scenarios_in_distance_interval(condition_path, min_distance, max_distance):
    verification_files = find_verification_files(condition_path)
    # Process each file
    attack_paths = []
    for file in verification_files:
        point_clouds = np.load(file)
        
        # Create the mask for the fourth element along the third axis
        mask = (point_clouds[ :, 3] == 6.0)
        
        # Use the mask to index into the array
        vehicle_points = point_clouds[mask]
        
        distance = calculate_minimum_distance_for_points(vehicle_points)

        if min_distance <= distance < max_distance:
            attack_paths.append(file[:-17] + ".npy")
    return attack_paths